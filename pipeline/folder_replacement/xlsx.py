"""XLSX native replacement with finite worksheet-cell text fitting."""

from __future__ import annotations

from collections.abc import Callable, Sequence
from copy import deepcopy
from dataclasses import dataclass
from io import BytesIO
from math import isfinite
from pathlib import Path
import posixpath
import re
from typing import cast
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo
import xml.etree.ElementTree as ElementTree

from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter,
    range_boundaries,
)
# skia-python does not publish PEP 561 stubs; this is the native measurement boundary.
import skia  # type: ignore[import-not-found]

from pipeline.bounded_text_layout import (
    BoundedTextBox,
    BoundedTextParagraph,
    BoundedTextRun,
    SourceTypefaceReference,
    noto_typefaces,
    replace_and_fit_text_box,
)
from pipeline.pptx_theme_fonts import PptxThemeFonts, resolve_theme_typefaces, theme_fonts_from_xml
from pipeline.folder_replacement.office_xml import (
    _namespace_bindings,
    _serialize_with_compatibility_bindings,
    _is_visible_office_text_element,
    replace_office_xml_text,
)
from pipeline.folder_replacement.failure_diagnostics import FailureContext
from pipeline.folder_replacement.common import NestedProgressReporter, is_office_bitmap_part
from pipeline.folder_replacement.bitmap import replace_bitmap_bytes
from pipeline.ocr import OcrProvider
from pipeline.ocr.image_preparation import DEFAULT_OCR_BACKGROUND
from pipeline.portable_fonts import static_noto_font
from pipeline.text_replacement import TextReplacementProvider, TextReplacementRequest, TextReplacementResult


_SPREADSHEET_NAMESPACE = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DRAWING_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DRAWING_MAIN_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/main"
_CHART_NAMESPACE = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_OFFICE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_RELATIONSHIPS_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_EMUS_PER_POINT = 12_700
_EMUS_PER_PIXEL = 9_525
_NS = {
    "x": _SPREADSHEET_NAMESPACE,
    "xdr": _DRAWING_NAMESPACE,
    "a": _DRAWING_MAIN_NAMESPACE,
    "r": _OFFICE_RELATIONSHIPS_NAMESPACE,
}
_NUMERIC_LOOKING_TEXT = re.compile(
    r"[+-]?(?:[0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.[0-9]+)?(?:[eE][+-]?[0-9]+)?%?"
)
XLSX_TRANSLATION_MODE_CHOICES = ("full", "fast")
_FAST_XLSX_MAX_WORKSHEET_ROWS = 1_000


@dataclass(frozen=True, slots=True)
class _FastXlsxCellSelection:
    cells: dict[str, frozenset[str]]
    skipped_worksheets: tuple[tuple[str, str], ...]


def replace_xlsx_file(
    source: Path,
    destination: Path,
    ocr: OcrProvider,
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    typeface: skia.Typeface,
    completed: Callable[[str], None],
    document_text_layout: str = "preserve-source-formatting",
    xlsx_translation_mode: str = "full",
    failure_context: FailureContext | None = None,
    nested_progress: NestedProgressReporter | None = None,
    diagnostics: list[dict[str, object]] | None = None,
) -> tuple[int, int, int]:
    """Replace XLSX cells without rewriting unrelated workbook package parts."""
    from pipeline.folder_replacement.processor import _replace_office_file

    if xlsx_translation_mode not in XLSX_TRANSLATION_MODE_CHOICES:
        raise ValueError(f"Unsupported XLSX translation mode: {xlsx_translation_mode!r}")
    if xlsx_translation_mode == "fast":
        with ZipFile(source) as archive:
            entries = [(entry, archive.read(entry.filename)) for entry in archive.infolist()]
        source_parts = {entry.filename: payload for entry, payload in entries}
        selection = _fast_xlsx_cell_selection(source_parts, ())
        _record_fast_mode_skipped_worksheets(diagnostics, selection)
        processing_provider: TextReplacementProvider = _XlsxTextSelectionProvider(
            _CurrentXlsxReplacementProvider(replacement, completed)
        )
        image_changes, image_regions = _replace_xlsx_fast_images(
            source_parts,
            ocr,
            replacement,
            source_language,
            target_language,
            typeface,
            completed,
            nested_progress,
        )
        changed_parts, native_items = _replace_xlsx_fast_parts(
            source_parts,
            processing_provider,
            source_language,
            target_language,
            measure_source_fonts=document_text_layout == "preserve-basic-layout-source-font",
            selection=selection,
            initial_changes=image_changes,
            failure_context=failure_context,
        )
        completed("chart cache synchronization")
        _write_xlsx_entries(destination, entries, changed_parts, failure_context)
        completed("package write")
        return native_items, image_regions, 0
    processing_provider = _XlsxTextSelectionProvider(replacement)
    if document_text_layout == "preserve-source-formatting":
        return _replace_office_file(
            source, destination, ocr, processing_provider, source_language, target_language, typeface, completed,
            failure_context=failure_context,
            nested_progress=nested_progress,
        )
    if document_text_layout not in {
        "preserve-basic-layout",
        "preserve-basic-layout-source-font",
    }:
        raise ValueError(f"Unsupported document text layout mode: {document_text_layout!r}")
    native_items, image_regions, retained_vectors = _replace_office_file(
        source,
        destination,
        ocr,
        processing_provider,
        source_language,
        target_language,
        typeface,
        completed,
        skip_native_xml_part=_is_custom_xlsx_part,
        failure_context=failure_context,
        nested_progress=nested_progress,
    )
    native_items += _replace_xlsx_cells(
        destination,
        processing_provider,
        source_language,
        target_language,
        # XLSX has no interoperable, package-level embedded-font path. Keep
        # the source reference while the shared fitter supplies its size.
        preserve_source_font_family=True,
        measure_source_fonts=document_text_layout == "preserve-basic-layout-source-font",
        failure_context=failure_context,
    )
    completed("native text layout")
    return native_items, image_regions, retained_vectors


def _is_custom_xlsx_part(name: str) -> bool:
    return (
        name == "xl/sharedStrings.xml"
        or name.startswith("xl/worksheets/")
        or name.startswith("xl/drawings/")
        or name.startswith("xl/tables/")
    )


def _replace_xlsx_cells(
    path: Path,
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    *,
    preserve_source_font_family: bool,
    measure_source_fonts: bool,
    failure_context: FailureContext | None = None,
    selected_cells: dict[str, frozenset[str]] | None = None,
) -> int:
    """Fit explicitly bounded cells and retain all unknown package parts byte-for-byte."""
    if failure_context is not None:
        failure_context.set_location(
            stage="xlsx_fitted_layout_read",
            container_kind="xlsx_document",
            operation="read",
        )
    with ZipFile(path) as archive:
        entries = [(entry, archive.read(entry.filename)) for entry in archive.infolist()]
    parts = {entry.filename: data for entry, data in entries}
    changed_parts, replacements = _replace_xlsx_cell_parts(
        parts,
        replacement,
        source_language,
        target_language,
        preserve_source_font_family=preserve_source_font_family,
        measure_source_fonts=measure_source_fonts,
        failure_context=failure_context,
        selected_cells=selected_cells,
    )
    if not changed_parts:
        return replacements
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for entry, data in entries:
            if failure_context is not None:
                failure_context.set_location(
                    stage="xlsx_fitted_layout_write",
                    container_kind="xlsx_package_part",
                    operation="write",
                    package_part=entry.filename,
                )
            archive.writestr(entry, changed_parts.get(entry.filename, data))
    if failure_context is not None:
        failure_context.set_location(
            stage="xlsx_fitted_layout_write",
            container_kind="xlsx_document",
            operation="write",
        )
    path.write_bytes(output.getvalue())
    return replacements


def replace_xlsx_bytes(
    data: bytes,
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    document_text_layout: str,
    nested_completed: Callable[[str], None] | None = None,
    xlsx_translation_mode: str = "full",
    fast_chart_formulae: tuple[str, ...] = (),
    ocr: OcrProvider | None = None,
    typeface: skia.Typeface | None = None,
    diagnostics: list[dict[str, object]] | None = None,
) -> tuple[bytes, int]:
    """Replace native XLSX text in memory for an enclosing OOXML package."""
    with ZipFile(BytesIO(data)) as archive:
        entries = [(entry, archive.read(entry.filename)) for entry in archive.infolist()]
    parts = {entry.filename: payload for entry, payload in entries}
    if xlsx_translation_mode not in XLSX_TRANSLATION_MODE_CHOICES:
        raise ValueError(f"Unsupported XLSX translation mode: {xlsx_translation_mode!r}")
    processing_provider: TextReplacementProvider = replacement
    if nested_completed is not None:
        processing_provider = _NestedXlsxReplacementProvider(replacement, nested_completed)
    processing_provider = _XlsxTextSelectionProvider(processing_provider)
    if xlsx_translation_mode == "fast":
        image_changes: dict[str, bytes] = {}
        if ocr is not None and typeface is not None:
            for name, payload in tuple(parts.items()):
                if is_office_bitmap_part(name):
                    parts[name], _regions = replace_bitmap_bytes(
                        payload,
                        ocr,
                        replacement,
                        source_language,
                        target_language,
                        typeface,
                        DEFAULT_OCR_BACKGROUND,
                        nested_completed,
                    )
                    image_changes[name] = parts[name]
        return _replace_xlsx_fast_bytes(
            entries,
            parts,
            processing_provider,
            source_language,
            target_language,
            document_text_layout,
            fast_chart_formulae,
            image_changes,
            diagnostics,
        )
    if document_text_layout == "preserve-source-formatting":
        changed_parts, replacements = _replace_xlsx_generic_xml_parts(
            parts, processing_provider, source_language, target_language
        )
    elif document_text_layout in {
        "preserve-basic-layout",
        "preserve-basic-layout-source-font",
    }:
        generic_changes, replacements = _replace_xlsx_generic_xml_parts(
            parts, processing_provider, source_language, target_language,
            skip_native_xml_part=_is_custom_xlsx_part,
        )
        processed_parts = {**parts, **generic_changes}
        changed_parts, cell_replacements = _replace_xlsx_cell_parts(
            processed_parts,
            processing_provider,
            source_language,
            target_language,
            preserve_source_font_family=True,
            measure_source_fonts=document_text_layout == "preserve-basic-layout-source-font",
        )
        changed_parts = {**generic_changes, **changed_parts}
        replacements += cell_replacements
    else:
        raise ValueError(f"Unsupported document text layout mode: {document_text_layout!r}")
    if not changed_parts:
        return data, replacements
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for entry, payload in entries:
            archive.writestr(entry, changed_parts.get(entry.filename, payload))
    return output.getvalue(), replacements


def xlsx_native_text_request_total(
    data: bytes,
    document_text_layout: str,
    xlsx_translation_mode: str = "full",
    fast_chart_formulae: tuple[str, ...] = (),
) -> int:
    """Return the exact number of replacement-provider calls an XLSX pass will make."""
    with ZipFile(BytesIO(data)) as archive:
        parts = {entry.filename: archive.read(entry.filename) for entry in archive.infolist()}
    if xlsx_translation_mode == "fast":
        selection = _fast_xlsx_cell_selection(parts, fast_chart_formulae)
        return (
            _fast_generic_text_request_total(parts)
            + _worksheet_text_request_count(parts, selection.cells)
            + sum(
                _drawing_text_request_count(payload)
                for name, payload in parts.items()
                if name.startswith("xl/drawings/") and name.endswith(".xml")
            )
            + 3 * sum(is_office_bitmap_part(name) for name in parts)
        )
    if document_text_layout == "preserve-source-formatting":
        return sum(
            _office_xml_text_request_count(payload)
            for name, payload in parts.items()
            if name.endswith(".xml")
        )
    if document_text_layout in {
        "preserve-basic-layout",
        "preserve-basic-layout-source-font",
    }:
        generic_requests = sum(
            _office_xml_text_request_count(payload)
            for name, payload in parts.items()
            if name.endswith(".xml") and not _is_custom_xlsx_part(name)
        )
        return generic_requests + _worksheet_text_request_count(parts) + sum(
            _drawing_text_request_count(payload)
            for name, payload in parts.items()
            if name.startswith("xl/drawings/") and name.endswith(".xml")
        )
    raise ValueError(f"Unsupported document text layout mode: {document_text_layout!r}")


def _office_xml_text_request_count(data: bytes) -> int:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return 0
    return sum(
        element.text is not None
        and _is_visible_office_text_element(element.tag)
        and _is_xlsx_translation_candidate(element.text)
        for element in root.iter()
    )


def _worksheet_text_request_count(
    parts: dict[str, bytes], selected_cells: dict[str, frozenset[str]] | None = None
) -> int:
    shared_strings = _shared_strings(parts.get("xl/sharedStrings.xml"))
    table_headers = _table_header_cells(parts, selected_cells)
    requests = 0
    for name, data in parts.items():
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        if selected_cells is not None and name not in selected_cells:
            continue
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError:
            continue
        for cell in root.findall(".//x:sheetData/x:row/x:c", _NS):
            if selected_cells is not None and cell.get("r") not in selected_cells.get(name, frozenset()):
                continue
            if cell.get("r") in table_headers.get(name, frozenset()):
                continue
            text = _cell_text(cell, shared_strings)
            if (
                text is not None
                and text.strip()
                and cell.find("x:f", _NS) is None
                and _is_xlsx_translation_candidate(text)
            ):
                requests += 1
    return requests


def _drawing_text_request_count(data: bytes) -> int:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return 0
    parents = {child: parent for parent in root.iter() for child in parent}
    fitted_text: set[ElementTree.Element] = set()
    requests = 0
    for body in root.findall(".//xdr:sp/xdr:txBody", _NS):
        shape = parents.get(body)
        extent = None if shape is None else shape.find("xdr:spPr/a:xfrm/a:ext", _NS)
        if extent is None:
            continue
        try:
            width, height = int(extent.get("cx", "0")), int(extent.get("cy", "0"))
        except ValueError:
            continue
        paragraph_text = tuple(
            "".join(
                "".join(text.text or "" for text in run.iter(_a_tag("t")))
                for run in paragraph.findall("a:r", _NS)
            )
            for paragraph in body.findall("a:p", _NS)
        )
        if width <= 0 or height <= 0 or not any(text.strip() for text in paragraph_text):
            continue
        requests += sum(_is_xlsx_translation_candidate(text) for text in paragraph_text)
        fitted_text.update(body.iter(_a_tag("t")))
    return requests + sum(
        element not in fitted_text
        and element.text is not None
        and _is_xlsx_translation_candidate(element.text)
        for element in root.iter(_a_tag("t"))
    )


def _replace_xlsx_generic_xml_parts(
    parts: dict[str, bytes],
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    *,
    skip_native_xml_part: Callable[[str], bool] | None = None,
    include_native_xml_part: Callable[[str], bool] | None = None,
) -> tuple[dict[str, bytes], int]:
    changed: dict[str, bytes] = {}
    replacements = 0
    for name, payload in parts.items():
        if (
            not name.endswith(".xml")
            or (skip_native_xml_part is not None and skip_native_xml_part(name))
            or (include_native_xml_part is not None and not include_native_xml_part(name))
        ):
            continue
        _set_nested_xlsx_part(replacement, name)
        updated, count = replace_office_xml_text(
            payload, replacement, source_language, target_language
        )
        if count:
            changed[name] = updated
            replacements += count
    return changed, replacements


def _replace_xlsx_fast_bytes(
    entries: Sequence[tuple[ZipInfo, bytes]],
    parts: dict[str, bytes],
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    document_text_layout: str,
    fast_chart_formulae: tuple[str, ...],
    initial_changes: dict[str, bytes],
    diagnostics: list[dict[str, object]] | None,
) -> tuple[bytes, int]:
    """Apply the bounded fast XLSX selection without touching unrelated sheets."""
    selection = _fast_xlsx_cell_selection(parts, fast_chart_formulae)
    _record_fast_mode_skipped_worksheets(diagnostics, selection)
    changed_parts, replacements = _replace_xlsx_fast_parts(
        parts,
        replacement,
        source_language,
        target_language,
        measure_source_fonts=document_text_layout == "preserve-basic-layout-source-font",
        selection=selection,
        initial_changes=initial_changes,
    )
    return _xlsx_entries_bytes(entries, changed_parts), replacements


def _replace_xlsx_fast_parts(
    parts: dict[str, bytes],
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    *,
    measure_source_fonts: bool,
    selection: _FastXlsxCellSelection,
    initial_changes: dict[str, bytes],
    failure_context: FailureContext | None = None,
) -> tuple[dict[str, bytes], int]:
    generic_changes, replacements = _replace_xlsx_generic_xml_parts(
        parts,
        replacement,
        source_language,
        target_language,
        include_native_xml_part=_is_fast_xlsx_generic_part,
    )
    processed_parts = {**parts, **generic_changes}
    cell_changes, cell_replacements = _replace_xlsx_cell_parts(
        processed_parts,
        replacement,
        source_language,
        target_language,
        preserve_source_font_family=True,
        measure_source_fonts=measure_source_fonts,
        failure_context=failure_context,
        selected_cells=selection.cells,
    )
    changed_parts = {**initial_changes, **generic_changes, **cell_changes}
    replacements += cell_replacements
    chart_cache_changes = _update_fast_xlsx_chart_caches({**parts, **changed_parts})
    changed_parts.update(chart_cache_changes)
    return changed_parts, replacements


def _replace_xlsx_fast_images(
    parts: dict[str, bytes],
    ocr: OcrProvider,
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    typeface: skia.Typeface,
    completed: Callable[[str], None],
    nested_progress: NestedProgressReporter | None,
) -> tuple[dict[str, bytes], int]:
    changes: dict[str, bytes] = {}
    image_regions = 0
    for name, payload in parts.items():
        if not is_office_bitmap_part(name):
            continue
        if nested_progress is not None:
            nested_progress.start_nested(name, 3, "stage")

        def image_stage_completed(label: str, *, image_name: str = name) -> None:
            completed(f"{image_name} {label}")
            if nested_progress is not None:
                nested_progress.advance_nested(label)

        try:
            changed, regions = replace_bitmap_bytes(
                payload,
                ocr,
                replacement,
                source_language,
                target_language,
                typeface,
                DEFAULT_OCR_BACKGROUND,
                image_stage_completed,
            )
        finally:
            if nested_progress is not None:
                nested_progress.clear_nested()
        changes[name] = changed
        image_regions += regions
    return changes, image_regions


def _write_xlsx_entries(
    destination: Path,
    entries: Sequence[tuple[ZipInfo, bytes]],
    changed_parts: dict[str, bytes],
    failure_context: FailureContext | None,
) -> None:
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for entry, payload in entries:
            if failure_context is not None:
                failure_context.set_location(
                    stage="xlsx_fast_package_write",
                    container_kind="xlsx_package_part",
                    operation="write",
                    package_part=entry.filename,
                )
            archive.writestr(entry, changed_parts.get(entry.filename, payload))
    if failure_context is not None:
        failure_context.set_location(
            stage="xlsx_fast_package_write",
            container_kind="xlsx_document",
            operation="write",
        )
    destination.write_bytes(output.getvalue())


def _xlsx_entries_bytes(
    entries: Sequence[tuple[ZipInfo, bytes]], changed_parts: dict[str, bytes]
) -> bytes:
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for entry, payload in entries:
            archive.writestr(entry, changed_parts.get(entry.filename, payload))
    return output.getvalue()


def _is_fast_xlsx_generic_part(name: str) -> bool:
    return (
        (name.startswith("xl/comments") and name.endswith(".xml"))
        or (name.startswith("xl/charts/") and name.endswith(".xml"))
    )


def _fast_generic_text_request_total(parts: dict[str, bytes]) -> int:
    return sum(
        _office_xml_text_request_count(payload)
        for name, payload in parts.items()
        if _is_fast_xlsx_generic_part(name)
    )


def _fast_xlsx_cell_selection(
    parts: dict[str, bytes], fast_chart_formulae: tuple[str, ...]
) -> _FastXlsxCellSelection:
    formulae = fast_chart_formulae or _xlsx_chart_formulae(parts)
    ranges = tuple(
        parsed for formula in formulae if (parsed := _xlsx_chart_formula_range(formula)) is not None
    )
    selected: dict[str, set[str]] = {}
    for sheet, first_column, first_row, last_column, last_row in ranges:
        selected.setdefault(sheet, set()).update(
            f"{get_column_letter(column)}{row}"
            for column in range(first_column, last_column + 1)
            for row in range(first_row, last_row + 1)
        )
    sheet_parts = _xlsx_sheet_parts(parts)
    shared_strings = _shared_strings(parts.get("xl/sharedStrings.xml"))
    skipped_worksheets: list[tuple[str, str]] = []
    for sheet_name, part_name in sheet_parts.items():
        data = parts.get(part_name)
        if data is None:
            continue
        if _worksheet_used_row_count(data) > _FAST_XLSX_MAX_WORKSHEET_ROWS:
            if sheet_name not in selected:
                skipped_worksheets.append((sheet_name, part_name))
            continue
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError:
            continue
        selected.setdefault(sheet_name, set()).update(
            reference
            for cell in root.findall(".//x:sheetData/x:row/x:c", _NS)
            if (reference := cell.get("r")) is not None
        )
    for sheet, first_column, first_row, last_column, last_row in ranges:
        referenced_part_name = sheet_parts.get(sheet)
        data = None if referenced_part_name is None else parts.get(referenced_part_name)
        if data is None:
            continue
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError:
            continue
        heading_row = _fast_heading_row(
            root, shared_strings, first_column, first_row, last_column, last_row
        )
        if heading_row is not None:
            selected.setdefault(sheet, set()).update(
                f"{get_column_letter(column)}{heading_row}"
                for column in range(first_column, last_column + 1)
            )
    return _FastXlsxCellSelection(
        {
            part_name: frozenset(selected.get(sheet_name, set()))
            for sheet_name, part_name in sheet_parts.items()
            if sheet_name in selected
        },
        tuple(skipped_worksheets),
    )


def _worksheet_used_row_count(data: bytes) -> int:
    """Return the stored-cell high-water mark, stopping once fast mode rejects it."""
    highest_row = 0
    try:
        for _event, row in ElementTree.iterparse(BytesIO(data), events=("end",)):
            if row.tag != _tag("row"):
                continue
            if not row.findall("x:c", _NS):
                row.clear()
                continue
            try:
                row_number = int(row.get("r", "0"))
            except ValueError:
                row_number = 0
            if row_number <= 0:
                for cell in row.findall("x:c", _NS):
                    reference = cell.get("r")
                    if reference is None:
                        continue
                    try:
                        _column, row_number = coordinate_from_string(reference)
                    except ValueError:
                        continue
                    break
            highest_row = max(highest_row, row_number)
            if highest_row > _FAST_XLSX_MAX_WORKSHEET_ROWS:
                return highest_row
            row.clear()
    except ElementTree.ParseError:
        return _FAST_XLSX_MAX_WORKSHEET_ROWS + 1
    return highest_row


def _record_fast_mode_skipped_worksheets(
    diagnostics: list[dict[str, object]] | None, selection: _FastXlsxCellSelection
) -> None:
    if diagnostics is None:
        return
    diagnostics.extend(
        {
            "kind": "skipped",
            "reason_code": "xlsx_fast_mode_worksheet_skipped",
            "container_kind": "xlsx_worksheet",
            "worksheet_name": worksheet_name,
            "location": {"package_part": part_name},
        }
        for worksheet_name, part_name in selection.skipped_worksheets
    )


def _xlsx_chart_formulae(parts: dict[str, bytes]) -> tuple[str, ...]:
    formulae: list[str] = []
    for name, data in parts.items():
        if not name.startswith("xl/charts/") or not name.endswith(".xml"):
            continue
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError:
            continue
        for reference in (*root.findall(f".//{{{_CHART_NAMESPACE}}}strRef"), *root.findall(f".//{{{_CHART_NAMESPACE}}}multiLvlStrRef")):
            formula = reference.findtext(f"{{{_CHART_NAMESPACE}}}f")
            if formula is not None:
                formulae.append(formula)
    return tuple(formulae)


def _xlsx_chart_formula_range(formula: str) -> tuple[str, int, int, int, int] | None:
    match = re.fullmatch(
        r"(?:(?P<quoted>'(?:[^']|'')+')|(?P<plain>[A-Za-z_][A-Za-z0-9_ ]*))!"
        r"\$?(?P<first_column>[A-Z]{1,3})\$?(?P<first_row>[1-9][0-9]*)"
        r"(?::\$?(?P<last_column>[A-Z]{1,3})\$?(?P<last_row>[1-9][0-9]*))?",
        formula,
    )
    if match is None:
        return None
    sheet = match.group("plain") or match.group("quoted")[1:-1].replace("''", "'")
    first_column = column_index_from_string(match.group("first_column"))
    first_row = int(match.group("first_row"))
    last_column = column_index_from_string(match.group("last_column") or match.group("first_column"))
    last_row = int(match.group("last_row") or match.group("first_row"))
    if last_column < first_column or last_row < first_row:
        return None
    return sheet, first_column, first_row, last_column, last_row


def _xlsx_sheet_parts(parts: dict[str, bytes]) -> dict[str, str]:
    try:
        workbook = ElementTree.fromstring(parts["xl/workbook.xml"])
        relationships = ElementTree.fromstring(parts["xl/_rels/workbook.xml.rels"])
    except (KeyError, ElementTree.ParseError):
        return {}
    relationship_targets = {
        relationship.get("Id"): _resolve_part_target("xl/workbook.xml", relationship.get("Target", ""))
        for relationship in relationships.findall(f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship")
        if relationship.get("TargetMode") != "External" and relationship.get("Id") is not None
    }
    result: dict[str, str] = {}
    for sheet in workbook.findall(".//x:sheet", _NS):
        name = sheet.get("name")
        relationship_id = sheet.get(f"{{{_OFFICE_RELATIONSHIPS_NAMESPACE}}}id")
        part_name = None if relationship_id is None else relationship_targets.get(relationship_id)
        if name is not None and part_name in parts:
            result[name] = part_name
    return result


def _fast_heading_row(
    worksheet: ElementTree.Element,
    shared_strings: tuple[str, ...],
    first_column: int,
    first_row: int,
    last_column: int,
    last_row: int,
) -> int | None:
    rows = [first_row - 1] if first_row > 1 else []
    rows.extend(range(first_row, last_row + 1))
    for row_number in rows:
        for column in range(first_column, last_column + 1):
            reference = f"{get_column_letter(column)}{row_number}"
            cell = next(
                (item for item in worksheet.findall(".//x:sheetData/x:row/x:c", _NS) if item.get("r") == reference),
                None,
            )
            if cell is not None and (_cell_text(cell, shared_strings) or "").strip():
                return row_number
    return None


def _update_fast_xlsx_chart_caches(parts: dict[str, bytes]) -> dict[str, bytes]:
    values = _xlsx_string_cell_values(parts, _chart_string_cell_references(parts))
    if values is None:
        return {}
    changed: dict[str, bytes] = {}
    for name, data in parts.items():
        if not name.startswith("xl/charts/") or not name.endswith(".xml"):
            continue
        try:
            root = ElementTree.fromstring(data)
        except ElementTree.ParseError:
            continue
        updated = False
        for reference in root.findall(f".//{{{_CHART_NAMESPACE}}}strRef"):
            cache = reference.find(f"{{{_CHART_NAMESPACE}}}strCache")
            updated = _update_fast_string_cache(reference, cache, values) or updated
        for reference in root.findall(f".//{{{_CHART_NAMESPACE}}}multiLvlStrRef"):
            cache = reference.find(f"{{{_CHART_NAMESPACE}}}multiLvlStrCache")
            updated = _update_fast_multi_level_cache(reference, cache, values) or updated
        if updated:
            changed[name] = _serialize_with_compatibility_bindings(root, _namespace_bindings(data))
    return changed


def _update_xlsx_fast_chart_caches_file(path: Path) -> None:
    with ZipFile(path) as archive:
        entries = [(entry, archive.read(entry.filename)) for entry in archive.infolist()]
    parts = {entry.filename: payload for entry, payload in entries}
    changes = _update_fast_xlsx_chart_caches(parts)
    if not changes:
        return
    with ZipFile(path, "w", ZIP_DEFLATED) as archive:
        for entry, payload in entries:
            archive.writestr(entry, changes.get(entry.filename, payload))


def _update_fast_string_cache(
    reference: ElementTree.Element,
    cache: ElementTree.Element | None,
    values: dict[tuple[str, str], str],
) -> bool:
    formula = reference.findtext(f"{{{_CHART_NAMESPACE}}}f")
    parsed = None if formula is None else _xlsx_chart_formula_range(formula)
    if cache is None or parsed is None:
        return False
    sheet, first_column, first_row, last_column, last_row = parsed
    cells = tuple(
        (sheet, f"{get_column_letter(column)}{row}")
        for row in range(first_row, last_row + 1)
        for column in range(first_column, last_column + 1)
    )
    points = cache.findall(f"{{{_CHART_NAMESPACE}}}pt")
    updates: list[tuple[ElementTree.Element, str]] = []
    for point in points:
        try:
            index = int(point.get("idx", ""))
        except ValueError:
            return False
        value = None if not 0 <= index < len(cells) else values.get(cells[index])
        output = point.find(f"{{{_CHART_NAMESPACE}}}v")
        if value is None or output is None:
            return False
        updates.append((output, value))
    for output, value in updates:
        output.text = value
    return bool(updates)


def _update_fast_multi_level_cache(
    reference: ElementTree.Element,
    cache: ElementTree.Element | None,
    values: dict[tuple[str, str], str],
) -> bool:
    formula = reference.findtext(f"{{{_CHART_NAMESPACE}}}f")
    parsed = None if formula is None else _xlsx_chart_formula_range(formula)
    if cache is None or parsed is None:
        return False
    sheet, first_column, first_row, last_column, last_row = parsed
    levels = cache.findall(f"{{{_CHART_NAMESPACE}}}lvl")
    if len(levels) != last_column - first_column + 1:
        return False
    updates: list[tuple[ElementTree.Element, str]] = []
    for offset, level in enumerate(levels):
        for point in level.findall(f"{{{_CHART_NAMESPACE}}}pt"):
            try:
                row = first_row + int(point.get("idx", ""))
            except ValueError:
                return False
            value = values.get((sheet, f"{get_column_letter(first_column + offset)}{row}"))
            output = point.find(f"{{{_CHART_NAMESPACE}}}v")
            if row > last_row or value is None or output is None:
                return False
            updates.append((output, value))
    for output, value in updates:
        output.text = value
    return bool(updates)


def _chart_string_cell_references(parts: dict[str, bytes]) -> dict[str, frozenset[str]]:
    references: dict[str, set[str]] = {}
    for formula in _xlsx_chart_formulae(parts):
        parsed = _xlsx_chart_formula_range(formula)
        if parsed is None:
            continue
        sheet, first_column, first_row, last_column, last_row = parsed
        references.setdefault(sheet, set()).update(
            f"{get_column_letter(column)}{row}"
            for column in range(first_column, last_column + 1)
            for row in range(first_row, last_row + 1)
        )
    return {sheet: frozenset(cells) for sheet, cells in references.items()}


def _xlsx_string_cell_values(
    parts: dict[str, bytes], requested_cells: dict[str, frozenset[str]] | None = None
) -> dict[tuple[str, str], str] | None:
    sheet_parts = _xlsx_sheet_parts(parts)
    if not sheet_parts:
        return None
    shared_strings = _shared_strings(parts.get("xl/sharedStrings.xml"))
    values: dict[tuple[str, str], str] = {}
    for sheet_name, part_name in sheet_parts.items():
        required = None if requested_cells is None else requested_cells.get(sheet_name)
        if requested_cells is not None and not required:
            continue
        try:
            root = ElementTree.fromstring(parts[part_name])
        except ElementTree.ParseError:
            return None
        for cell in root.findall(".//x:sheetData/x:row/x:c", _NS):
            reference = cell.get("r")
            value = _cell_text(cell, shared_strings)
            if reference is not None and value is not None and (required is None or reference in required):
                values[sheet_name, reference] = value
    return values


def _replace_xlsx_cell_parts(
    parts: dict[str, bytes],
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    *,
    preserve_source_font_family: bool,
    measure_source_fonts: bool,
    failure_context: FailureContext | None = None,
    selected_cells: dict[str, frozenset[str]] | None = None,
) -> tuple[dict[str, bytes], int]:
    shared_strings = _shared_strings(parts.get("xl/sharedStrings.xml"))
    theme = _workbook_theme(parts)
    styles = _Styles(parts.get("xl/styles.xml"), theme)
    typefaces = noto_typefaces()
    table_headers = _table_header_cells(parts, selected_cells)
    replacements = 0
    changed_parts: dict[str, bytes] = {}
    for name, data in parts.items():
        if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
            continue
        if selected_cells is not None and name not in selected_cells:
            continue
        if failure_context is not None:
            failure_context.set_location(
                stage="xlsx_fitted_layout",
                container_kind="xlsx_worksheet",
                operation="text_replacement",
                package_part=name,
            )
        _set_nested_xlsx_part(replacement, name)
        updated, count = _replace_worksheet(
            data,
            shared_strings,
            styles,
            replacement,
            source_language,
            target_language,
            typefaces,
            preserve_source_font_family,
            measure_source_fonts,
            table_headers.get(name, frozenset()),
            None if selected_cells is None else selected_cells.get(name, frozenset()),
        )
        changed_parts[name] = updated
        replacements += count
    for name, data in parts.items():
        if not name.startswith("xl/drawings/") or not name.endswith(".xml"):
            continue
        if failure_context is not None:
            failure_context.set_location(
                stage="xlsx_fitted_layout",
                container_kind="xlsx_drawing",
                operation="text_replacement",
                package_part=name,
            )
        _set_nested_xlsx_part(replacement, name)
        updated, count = _replace_drawing(
            data, replacement, source_language, target_language, typefaces,
            preserve_source_font_family, measure_source_fonts, theme,
        )
        changed_parts[name] = updated
        replacements += count
    if styles.changed:
        changed_parts["xl/styles.xml"] = styles.serialize()
    return changed_parts, replacements


class _NestedXlsxReplacementProvider:
    """Report the active XLSX replacement request without retaining its text."""

    def __init__(self, provider: TextReplacementProvider, completed: Callable[[str], None]) -> None:
        self._provider = provider
        self._completed = completed
        self._part = "embedded workbook"
        self._count = 0

    def set_part(self, part: str) -> None:
        self._part = part

    def replace(self, request: TextReplacementRequest) -> TextReplacementResult:
        self._count += 1
        self._completed(f"{self._part} replacement {self._count}")
        return self._provider.replace(request)


class _CurrentXlsxReplacementProvider:
    """Advance standalone fast-mode progress after an eligible replacement."""

    def __init__(self, provider: TextReplacementProvider, completed: Callable[[str], None]) -> None:
        self._provider = provider
        self._completed = completed
        self._part = "workbook"
        self._count = 0

    def set_part(self, part: str) -> None:
        self._part = part

    def replace(self, request: TextReplacementRequest) -> TextReplacementResult:
        result = self._provider.replace(request)
        self._count += 1
        self._completed(f"{self._part} replacement {self._count}")
        return result


class _XlsxTextSelectionProvider:
    """Skip values whose textual form is defined as non-translatable in XLSX."""

    def __init__(self, provider: TextReplacementProvider) -> None:
        self._provider = provider

    def set_part(self, part: str) -> None:
        if isinstance(self._provider, (_CurrentXlsxReplacementProvider, _NestedXlsxReplacementProvider)):
            self._provider.set_part(part)

    def replace(self, request: TextReplacementRequest) -> TextReplacementResult:
        if not request.is_filename and not _is_xlsx_translation_candidate(request.text):
            return TextReplacementResult(request.text, 0.0)
        return self._provider.replace(request)


def _set_nested_xlsx_part(provider: TextReplacementProvider, part: str) -> None:
    if isinstance(provider, (_CurrentXlsxReplacementProvider, _NestedXlsxReplacementProvider, _XlsxTextSelectionProvider)):
        provider.set_part(part)


def _is_xlsx_translation_candidate(text: str) -> bool:
    """Return whether a visible XLSX string is outside the numeric-only exclusion."""
    return _NUMERIC_LOOKING_TEXT.fullmatch(text) is None


def _replace_worksheet(
    data: bytes,
    shared_strings: tuple[str, ...],
    styles: "_Styles",
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
    typefaces: dict[str, skia.Typeface],
    preserve_source_font_family: bool,
    measure_source_fonts: bool,
    table_headers: frozenset[str],
    selected_cells: frozenset[str] | None = None,
) -> tuple[bytes, int]:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return data, 0
    column_widths = _column_widths(root)
    row_heights = _row_heights(root)
    merged_cells = _merged_cells(root)
    replaced = 0
    for cell in root.findall(".//x:sheetData/x:row/x:c", _NS):
        if selected_cells is not None and cell.get("r") not in selected_cells:
            continue
        if cell.get("r") in table_headers:
            # A structured-table header is an identifier referenced by table
            # formulas and metadata, rather than ordinary display text.
            continue
        text = _cell_text(cell, shared_strings)
        if (
            text is None
            or not text.strip()
            or cell.find("x:f", _NS) is not None
            or not _is_xlsx_translation_candidate(text)
        ):
            continue
        bounds = _cell_bounds(cell, column_widths, row_heights, merged_cells)
        if not styles.can_write_explicit_fit:
            bounds = None
        if bounds is None:
            _write_cell_text(
                cell,
                _replace_text(text, replacement, source_language, target_language),
            )
            replaced += 1
            continue
        source_run = styles.run_for(cell)
        text_box = BoundedTextBox(
            width_emu=bounds[0],
            height_emu=bounds[1],
            margin_left_emu=0,
            margin_top_emu=0,
            margin_right_emu=0,
            margin_bottom_emu=0,
            text_direction=None,
            paragraphs=(
                BoundedTextParagraph(
                    alignment=styles.alignment_for(cell),
                    space_before_points=None,
                    space_after_points=None,
                    line_spacing=None,
                    line_spacing_kind=None,
                    level=0,
                    margin_left_emu=None,
                    indent_emu=None,
                    bullet_kind=None,
                    bullet_marker=None,
                    empty_line_font_size_points=None,
                    runs=(replace_run_text(source_run, text),),
                ),
            ),
        )
        fitted = replace_and_fit_text_box(
            text_box,
            replacement,
            source_language,
            target_language,
            typefaces,
            preserve_source_font_family=preserve_source_font_family,
            measure_source_fonts=measure_source_fonts,
        )
        explicit_runs = fitted.text_box.paragraphs[0].runs
        explicit_run = explicit_runs[0]
        if len(explicit_runs) == 1:
            _write_cell_text(cell, explicit_run.text)
        else:
            _write_rich_cell_text(cell, explicit_runs)
        styles.apply_explicit_fit(cell, explicit_run)
        replaced += 1
    return _serialize_with_compatibility_bindings(root, _namespace_bindings(data)), replaced


def _table_header_cells(
    parts: dict[str, bytes], selected_parts: dict[str, frozenset[str]] | None = None
) -> dict[str, frozenset[str]]:
    """Return header-cell references that must agree with table metadata.

    A worksheet table's header row is duplicated in ``xl/tables/*.xml`` as
    structured-reference names.  Replacing only the visible cell text leaves
    those two representations inconsistent, which Excel repairs by rebuilding
    the table.  Keep both header forms unchanged; body cells remain ordinary
    worksheet text and are processed below.
    """
    headers_by_sheet: dict[str, frozenset[str]] = {}
    for sheet_name, sheet_data in parts.items():
        if not sheet_name.startswith("xl/worksheets/") or not sheet_name.endswith(".xml"):
            continue
        if selected_parts is not None and sheet_name not in selected_parts:
            continue
        try:
            sheet = ElementTree.fromstring(sheet_data)
        except ElementTree.ParseError:
            continue
        relationships = _part_relationships(parts, sheet_name)
        headers: set[str] = set()
        for table_part in sheet.findall(".//x:tableParts/x:tablePart", _NS):
            relationship_id = table_part.get(f"{{{_OFFICE_RELATIONSHIPS_NAMESPACE}}}id")
            target = None if relationship_id is None else relationships.get(relationship_id)
            if target is None:
                continue
            table_data = parts.get(_resolve_part_target(sheet_name, target))
            if table_data is None:
                continue
            try:
                table = ElementTree.fromstring(table_data)
                reference = table.get("ref")
                if reference is None:
                    continue
                min_column, min_row, max_column, _max_row = range_boundaries(reference)
            except (ElementTree.ParseError, ValueError):
                continue
            if min_column is None or min_row is None or max_column is None:
                continue
            headers.update(
                f"{get_column_letter(column)}{min_row}"
                for column in range(min_column, max_column + 1)
            )
        headers_by_sheet[sheet_name] = frozenset(headers)
    return headers_by_sheet


def _part_relationships(parts: dict[str, bytes], part_name: str) -> dict[str, str]:
    parent, basename = posixpath.split(part_name)
    relationships_name = f"{parent}/_rels/{basename}.rels"
    data = parts.get(relationships_name)
    if data is None:
        return {}
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return {}
    relationships: dict[str, str] = {}
    for relationship in root.findall(f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"):
        relationship_id = relationship.get("Id")
        target = relationship.get("Target")
        if relationship_id is not None and target is not None and relationship.get("TargetMode") != "External":
            relationships[relationship_id] = target
    return relationships


def _workbook_theme(parts: dict[str, bytes]) -> PptxThemeFonts | None:
    relationships_name = "xl/_rels/workbook.xml.rels"
    data = parts.get(relationships_name)
    if data is None:
        return None
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return None
    for relationship in root.findall(f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"):
        if relationship.get("TargetMode") == "External" or not (relationship.get("Type") or "").endswith("/theme"):
            continue
        target = relationship.get("Target")
        if target:
            return theme_fonts_from_xml(parts.get(_resolve_part_target("xl/workbook.xml", target)))
    return None


def _resolve_part_target(part_name: str, target: str) -> str:
    """Resolve a package-relative relationship target without filesystem access."""
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(part_name), target))


def _replace_drawing(
    data: bytes,
    provider: TextReplacementProvider,
    source_language: str,
    target_language: str,
    typefaces: dict[str, skia.Typeface],
    preserve_source_font_family: bool,
    measure_source_fonts: bool = False,
    theme: PptxThemeFonts | None = None,
) -> tuple[bytes, int]:
    """Fit SpreadsheetDrawing shape text against its ``a:xfrm/a:ext`` box."""
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return data, 0
    parents = {child: parent for parent in root.iter() for child in parent}
    fitted_text: set[ElementTree.Element] = set()
    count = 0
    for body in root.findall(".//xdr:sp/xdr:txBody", _NS):
        shape = parents.get(body)
        extent = None if shape is None else shape.find("xdr:spPr/a:xfrm/a:ext", _NS)
        if extent is None:
            continue
        try:
            width, height = int(extent.get("cx", "0")), int(extent.get("cy", "0"))
        except ValueError:
            continue
        paragraphs = tuple(_drawing_paragraph(paragraph, theme) for paragraph in body.findall("a:p", _NS))
        if width <= 0 or height <= 0 or not any(run.text.strip() for paragraph in paragraphs for run in paragraph.runs):
            continue
        body_properties = body.find("a:bodyPr", _NS)
        box = BoundedTextBox(
            width, height,
            _int_attribute(body_properties, "lIns", 91_440), _int_attribute(body_properties, "tIns", 45_720),
            _int_attribute(body_properties, "rIns", 91_440), _int_attribute(body_properties, "bIns", 45_720),
            None, paragraphs,
        )
        fitted = replace_and_fit_text_box(
            box, provider, source_language, target_language, typefaces,
            preserve_source_font_family=preserve_source_font_family,
            measure_source_fonts=measure_source_fonts,
        )
        for element in body.iter(_a_tag("t")):
            fitted_text.add(element)
        for destination, explicit in zip(body.findall("a:p", _NS), fitted.text_box.paragraphs, strict=True):
            _write_drawing_paragraph(destination, explicit)
        for element in body.iter(_a_tag("t")):
            fitted_text.add(element)
        count += sum(bool("".join(run.text for run in paragraph.runs).strip()) for paragraph in paragraphs)
    for element in root.iter(_a_tag("t")):
        if element not in fitted_text and element.text:
            element.text = _replace_text(element.text, provider, source_language, target_language)
            count += 1
    return _serialize_with_compatibility_bindings(root, _namespace_bindings(data)), count


def _drawing_paragraph(element: ElementTree.Element, theme: PptxThemeFonts | None) -> BoundedTextParagraph:
    properties = element.find("a:pPr", _NS)
    alignment = None if properties is None else properties.get("algn")
    runs = tuple(_drawing_run(run, theme) for run in element.findall("a:r", _NS))
    return BoundedTextParagraph(alignment or "left", None, None, None, None, 0, None, None,
                                None, None, None, runs)


def _drawing_run(element: ElementTree.Element, theme: PptxThemeFonts | None) -> BoundedTextRun:
    properties = element.find("a:rPr", _NS)
    references = _drawing_source_typefaces(properties, theme, "".join(item.text or "" for item in element.iter(_a_tag("t"))))
    family = next((item.original_family for item in references if item.script == "latin"), None)
    size = _finite_float(None if properties is None else properties.get("sz"))
    return BoundedTextRun(
        "".join(item.text or "" for item in element.iter(_a_tag("t"))), family,
        _font_classification(family), None if size is None else size / 100.0,
        properties is not None and properties.get("b") == "1", properties is not None and properties.get("i") == "1",
        "single" if properties is not None and properties.get("u") not in {None, "none"} else "none", None, references,
    )


def _drawing_source_typefaces(
    properties: ElementTree.Element | None, theme: PptxThemeFonts | None, text: str
) -> tuple[SourceTypefaceReference, ...]:
    if properties is None:
        return ()
    references: list[SourceTypefaceReference] = []
    for script, tag in (("latin", "latin"), ("eastAsian", "ea"), ("complex", "cs")):
        element = properties.find(_a_tag(tag))
        if element is not None and element.get("typeface"):
            references.append(SourceTypefaceReference(script, element.get("typeface")))
    return resolve_theme_typefaces(tuple(references), theme, text)


def _write_drawing_paragraph(element: ElementTree.Element, paragraph: BoundedTextParagraph) -> None:
    for child in tuple(element):
        if child.tag in {_a_tag("r"), _a_tag("br"), _a_tag("fld")}:
            element.remove(child)
    end_paragraph = element.find("a:endParaRPr", _NS)
    insertion_index = len(element) if end_paragraph is None else list(element).index(end_paragraph)
    for run in paragraph.runs:
        destination = ElementTree.Element(_a_tag("r"))
        properties = ElementTree.SubElement(destination, _a_tag("rPr"))
        references = run.source_typefaces or (SourceTypefaceReference("latin", run.font_family or "Noto Sans JP"),)
        for item in references:
            ElementTree.SubElement(properties, _a_tag({"latin": "latin", "eastAsian": "ea", "complex": "cs"}[item.script]), {"typeface": item.original_family or "Noto Sans JP"})
        properties.set("sz", str(max(100, round((run.font_size_points or 18.0) * 100))))
        if run.bold: properties.set("b", "1")
        if run.italic: properties.set("i", "1")
        if run.underline not in {None, "none"}: properties.set("u", "sng")
        ElementTree.SubElement(destination, _a_tag("t")).text = run.text
        element.insert(insertion_index, destination)
        insertion_index += 1


def _int_attribute(element: ElementTree.Element | None, name: str, default: int) -> int:
    try:
        return int(default if element is None else element.get(name, str(default)))
    except ValueError:
        return default


def _a_tag(local_name: str) -> str:
    return f"{{{_DRAWING_MAIN_NAMESPACE}}}{local_name}"


def replace_run_text(run: BoundedTextRun, text: str) -> BoundedTextRun:
    """Retain a source run's style while supplying its complete cell value."""
    return BoundedTextRun(
        text=text,
        font_family=run.font_family,
        font_classification=run.font_classification,
        font_size_points=run.font_size_points,
        bold=run.bold,
        italic=run.italic,
        underline=run.underline,
        baseline=run.baseline,
        source_typefaces=run.source_typefaces,
    )


def _replace_text(
    text: str,
    replacement: TextReplacementProvider,
    source_language: str,
    target_language: str,
) -> str:
    return replacement.replace(
        TextReplacementRequest(text, False, source_language, target_language)
    ).text


def _shared_strings(data: bytes | None) -> tuple[str, ...]:
    if data is None:
        return ()
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return ()
    return tuple("".join(element.text or "" for element in item.iter(_tag("t"))) for item in root)


def _cell_text(cell: ElementTree.Element, shared_strings: tuple[str, ...]) -> str | None:
    cell_type = cell.get("t")
    if cell_type == "s":
        value = cell.find("x:v", _NS)
        if value is None or value.text is None:
            return None
        try:
            return shared_strings[int(value.text)]
        except (IndexError, ValueError):
            return None
    if cell_type == "inlineStr":
        inline = cell.find("x:is", _NS)
        return None if inline is None else "".join(element.text or "" for element in inline.iter(_tag("t")))
    return None


def _write_cell_text(cell: ElementTree.Element, text: str) -> None:
    for child in tuple(cell):
        if child.tag in {_tag("v"), _tag("is")}:
            cell.remove(child)
    cell.set("t", "inlineStr")
    inline = ElementTree.SubElement(cell, _tag("is"))
    value = ElementTree.SubElement(inline, _tag("t"))
    value.text = text
    if text[:1].isspace() or text[-1:].isspace():
        value.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")


def _write_rich_cell_text(cell: ElementTree.Element, runs: tuple[BoundedTextRun, ...]) -> None:
    """Write portable fallback spans as SpreadsheetML rich-text runs."""
    for child in tuple(cell):
        if child.tag in {_tag("v"), _tag("is")}:
            cell.remove(child)
    cell.set("t", "inlineStr")
    inline = ElementTree.SubElement(cell, _tag("is"))
    for run in runs:
        destination = ElementTree.SubElement(inline, _tag("r"))
        properties = ElementTree.SubElement(destination, _tag("rPr"))
        family, _path = static_noto_font(run.font_classification, run.bold)
        ElementTree.SubElement(properties, _tag("rFont"), {"val": family})
        ElementTree.SubElement(properties, _tag("sz"), {"val": f"{run.font_size_points or 18.0:.4f}"})
        if run.bold:
            ElementTree.SubElement(properties, _tag("b"))
        if run.italic:
            ElementTree.SubElement(properties, _tag("i"))
        if run.underline not in {None, "none"}:
            ElementTree.SubElement(properties, _tag("u"))
        value = ElementTree.SubElement(destination, _tag("t"))
        value.text = run.text
        if run.text[:1].isspace() or run.text[-1:].isspace():
            value.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")


def _column_widths(root: ElementTree.Element) -> dict[int, int]:
    widths: dict[int, int] = {}
    for dimension in root.findall("x:cols/x:col", _NS):
        width = _finite_float(dimension.get("width"))
        if width is None:
            continue
        pixels = int(((256 * width + int(128 / 7)) / 256) * 7)
        if pixels <= 0:
            continue
        for column in range(int(dimension.get("min", "0")), int(dimension.get("max", "-1")) + 1):
            widths[column] = pixels * _EMUS_PER_PIXEL
    return widths


def _row_heights(root: ElementTree.Element) -> dict[int, int]:
    heights: dict[int, int] = {}
    for row in root.findall(".//x:sheetData/x:row", _NS):
        height = _finite_float(row.get("ht"))
        if height is not None and height > 0 and row.get("r") is not None:
            heights[int(row.get("r", "0"))] = int(round(height * _EMUS_PER_POINT))
    return heights


def _merged_cells(root: ElementTree.Element) -> tuple[tuple[int, int, int, int], ...]:
    merged: list[tuple[int, int, int, int]] = []
    for element in root.findall("x:mergeCells/x:mergeCell", _NS):
        reference = element.get("ref")
        if reference is None:
            continue
        min_column, min_row, max_column, max_row = range_boundaries(reference)
        if None in (min_column, min_row, max_column, max_row):
            continue
        merged.append(cast(tuple[int, int, int, int], (min_column, min_row, max_column, max_row)))
    return tuple(merged)


def _cell_bounds(
    cell: ElementTree.Element,
    column_widths: dict[int, int],
    row_heights: dict[int, int],
    merged_cells: tuple[tuple[int, int, int, int], ...],
) -> tuple[int, int] | None:
    reference = cell.get("r")
    if reference is None:
        return None
    column_name, row = coordinate_from_string(reference)
    column = column_index_from_string(column_name)
    min_column, min_row, max_column, max_row = column, row, column, row
    for candidate in merged_cells:
        if candidate[0] <= column <= candidate[2] and candidate[1] <= row <= candidate[3]:
            if (column, row) != candidate[:2]:
                return None
            min_column, min_row, max_column, max_row = candidate
            break
    widths = tuple(column_widths.get(index) for index in range(min_column, max_column + 1))
    heights = tuple(row_heights.get(index) for index in range(min_row, max_row + 1))
    if any(value is None for value in widths + heights):
        return None
    return (
        sum(value for value in widths if value is not None),
        sum(value for value in heights if value is not None),
    )


def _finite_float(value: str | None) -> float | None:
    if value is None:
        return None
    parsed = float(value)
    return parsed if isfinite(parsed) else None


class _Styles:
    """Mutable styles.xml subset used to attach explicit fitted cell formatting."""

    def __init__(self, data: bytes | None, theme: PptxThemeFonts | None = None) -> None:
        self._data = data
        self._root = ElementTree.fromstring(data) if data is not None else None
        self.changed = False
        self._theme = theme

    @property
    def can_write_explicit_fit(self) -> bool:
        return (
            self._root is not None
            and self._root.find(_tag("fonts")) is not None
            and self._root.find(_tag("cellXfs")) is not None
        )

    def run_for(self, cell: ElementTree.Element) -> BoundedTextRun:
        font = self._font_for(cell)
        scheme = None if font is None else font.find(_tag("scheme"))
        scheme_value = None if scheme is None else scheme.get("val")
        alias = {"major": "+mj-lt", "minor": "+mn-lt"}.get(scheme_value or "")
        references = () if alias is None or _font_name(font) else (
            SourceTypefaceReference("latin", scheme_value, self._theme.resolve(alias, "latin") if self._theme else None),
        )
        family = _font_name(font)
        return BoundedTextRun(
            text="",
            font_family=family,
            font_classification=_font_classification(family),
            font_size_points=_font_size(font),
            bold=font is not None and font.find(_tag("b")) is not None,
            italic=font is not None and font.find(_tag("i")) is not None,
            underline="single" if font is not None and font.find(_tag("u")) is not None else "none",
            baseline=_font_baseline(font),
            source_typefaces=references,
        )

    def alignment_for(self, cell: ElementTree.Element) -> str:
        xf = self._xf_for(cell)
        alignment = None if xf is None else xf.find(_tag("alignment"))
        return {"center": "center", "right": "right", "justify": "justify"}.get(
            alignment.get("horizontal", "") if alignment is not None else "", "left"
        )

    def apply_explicit_fit(self, cell: ElementTree.Element, run: BoundedTextRun) -> None:
        if self._root is None:
            return
        fonts = self._root.find(_tag("fonts"))
        cell_xfs = self._root.find(_tag("cellXfs"))
        source_xf = self._xf_for(cell)
        source_font = self._font_for(cell)
        if fonts is None or cell_xfs is None or source_xf is None:
            return
        font = deepcopy(source_font) if source_font is not None else ElementTree.Element(_tag("font"))
        if not run.source_typefaces:
            _set_font_value(font, "name", "val", run.font_family or "Noto Sans JP")
        _set_font_value(font, "sz", "val", f"{run.font_size_points or 18.0:.4f}")
        fonts.append(font)
        fonts.set("count", str(len(fonts)))
        xf = deepcopy(source_xf)
        xf.set("fontId", str(len(fonts) - 1))
        xf.set("applyFont", "1")
        alignment = xf.find(_tag("alignment"))
        if alignment is None:
            alignment = ElementTree.Element(_tag("alignment"))
            protection = xf.find(_tag("protection"))
            if protection is None:
                xf.append(alignment)
            else:
                xf.insert(list(xf).index(protection), alignment)
        alignment.set("wrapText", "1")
        alignment.set("shrinkToFit", "0")
        cell_xfs.append(xf)
        cell_xfs.set("count", str(len(cell_xfs)))
        cell.set("s", str(len(cell_xfs) - 1))
        self.changed = True

    def serialize(self) -> bytes:
        if self._root is None or self._data is None:
            return self._data or b""
        return _serialize_with_compatibility_bindings(self._root, _namespace_bindings(self._data))

    def _xf_for(self, cell: ElementTree.Element) -> ElementTree.Element | None:
        if self._root is None:
            return None
        cell_xfs = self._root.find(_tag("cellXfs"))
        if cell_xfs is None:
            return None
        index = int(cell.get("s", "0"))
        return cell_xfs[index] if 0 <= index < len(cell_xfs) else None

    def _font_for(self, cell: ElementTree.Element) -> ElementTree.Element | None:
        if self._root is None:
            return None
        xf = self._xf_for(cell)
        fonts = self._root.find(_tag("fonts"))
        if xf is None or fonts is None:
            return None
        index = int(xf.get("fontId", "0"))
        return fonts[index] if 0 <= index < len(fonts) else None


def _font_name(font: ElementTree.Element | None) -> str | None:
    name = None if font is None else font.find(_tag("name"))
    return None if name is None else name.get("val")


def _font_size(font: ElementTree.Element | None) -> float | None:
    size = None if font is None else font.find(_tag("sz"))
    return _finite_float(None if size is None else size.get("val"))


def _font_baseline(font: ElementTree.Element | None) -> int | None:
    value = None if font is None else font.find(_tag("vertAlign"))
    if value is None:
        return None
    return {"superscript": 30_000, "subscript": -25_000}.get(value.get("val", ""))


def _set_font_value(font: ElementTree.Element, name: str, attribute: str, value: str) -> None:
    child = font.find(_tag(name))
    if child is None:
        child = ElementTree.SubElement(font, _tag(name))
    child.set(attribute, value)


def _font_classification(family: str | None) -> str:
    normalized = (family or "").lower()
    if any(marker in normalized for marker in ("mono", "code", "courier", "fixed")):
        return "fixed-width"
    if any(marker in normalized for marker in ("serif", "mincho", "ming", "song")):
        return "serif"
    return "sans-serif"


def _tag(local_name: str) -> str:
    return f"{{{_SPREADSHEET_NAMESPACE}}}{local_name}"


__all__ = ["replace_xlsx_bytes", "replace_xlsx_file", "xlsx_native_text_request_total"]
