#!/usr/bin/env python3
"""Synthetic regression tests for XLSX folder replacement."""

from __future__ import annotations

from contextlib import redirect_stderr, redirect_stdout
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
from tempfile import TemporaryDirectory
from typing import cast
import unittest
from unittest.mock import patch
from uuid import UUID
import xml.etree.ElementTree as ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from PIL import Image
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.worksheet.table import Table
from pptx import Presentation
from pptx.enum.text import MSO_ANCHOR, MSO_AUTO_SIZE
from pptx.oxml import parse_xml
from pptx.oxml.ns import qn
from pptx.util import Inches, Pt
# pypdf does not publish PEP 561 metadata for its generic object model.
from pypdf import PdfReader, PdfWriter
from pypdf.generic import ArrayObject, ByteStringObject, ContentStream, DecodedStreamObject, DictionaryObject, FloatObject, NameObject, NumberObject, TextStringObject
# skia-python does not publish PEP 561 stubs; this is the native rendering boundary.
import skia  # type: ignore[import-not-found]

from pipeline.folder_replacement import (
    FolderReplacementResult,
    parse_include_patterns,
    replace_input_folder,
)
from pipeline.folder_replacement.processor import ProgressFactory, ProgressReporter
from pipeline.folder_replacement.docx import _docx_ocr_backgrounds
from pipeline.folder_replacement.pptx import _pptx_ocr_backgrounds
from pipeline.folder_replacement.pdf import (
    _PdfPaintSpan,
    _PdfShownText,
    _PdfVisualRegion,
    _PdfReplacementSerializationError,
    _pdf_apply_paint_span_bullet_overrides,
    _pdf_apply_legacy_bullet_override,
    _pdf_content_has_annotations,
    _pdf_decode_composite_bytes,
    _pdf_expansion_geometry_is_known,
    _pdf_fitted_region_operations,
    _pdf_is_candidate_bullet_error,
    _pdf_text_advance,
)
from pipeline.folder_replacement.xlsx import (
    _replace_drawing,
    replace_xlsx_bytes,
    xlsx_native_text_request_total,
)
from pipeline.folder_replacement.docx import _validate_docx_embedded_fonts
from pipeline.bounded_text_layout import (
    BoundedTextBox,
    BoundedTextParagraph,
    BoundedTextRun,
    PortableTextUnsupportedError,
    noto_typefaces,
)
from pipeline.portable_bullet_overrides import LegacyBulletOverride
from pipeline.portable_fonts import static_noto_bytes
from pipeline.ocr import BoundingPolygon, OcrRequest, OcrResult, OcrText, PixelPoint
from pipeline.ocr.provider import LocalContractTestSkip
from pipeline.text_replacement import (
    TextReplacementProvider,
    TextReplacementRequest,
    TextReplacementResult,
)
from pipeline.text_replacement_plugins.character_mask import CharacterMaskProvider



from folder_replacement_test_support import (
    FolderReplacementTestCase,
    _CountingOcrProvider,
    _EmptyOcrProvider,
    _FailingOcrProvider,
    _LowConfidenceOcrProvider,
    _RecordedProgress,
    _RecordingReplacementProvider,
    _VectorOutlineOcrProvider,
    _synthetic_pdf_visual_region,
)

class FolderReplacementXlsxTests(FolderReplacementTestCase):
    # Verifies FR-2026-09-04-01.
    def test_fast_mode_translates_all_cells_on_worksheets_with_at_most_1000_used_rows(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "workbook.xlsx"
            workbook = Workbook()
            small_sheet = workbook.active
            assert small_sheet is not None
            small_sheet.title = "Small"
            small_sheet["A1"] = "Small first"
            small_sheet["A500"] = "Small last"
            large_sheet = workbook.create_sheet("Large")
            large_sheet["A1"] = "Large first"
            large_sheet["A1001"] = "Large last"
            workbook.save(path)
            with ZipFile(path) as archive:
                source_large_sheet = archive.read("xl/worksheets/sheet2.xml")

            request_total = xlsx_native_text_request_total(
                path.read_bytes(), "preserve-source-formatting", xlsx_translation_mode="fast"
            )
            provider = _RecordingReplacementProvider(replacement_text="Translated")
            translated, replacements = replace_xlsx_bytes(
                path.read_bytes(),
                provider,
                "en",
                "fr",
                "preserve-source-formatting",
                xlsx_translation_mode="fast",
            )

        self.assertEqual(2, request_total)
        self.assertEqual(2, replacements)
        self.assertEqual(["Small first", "Small last"], [request.text for request in provider.requests])
        output = load_workbook(BytesIO(translated))
        self.assertEqual("Translated", output["Small"]["A1"].value)
        self.assertEqual("Translated", output["Small"]["A500"].value)
        self.assertEqual("Large first", output["Large"]["A1"].value)
        self.assertEqual("Large last", output["Large"]["A1001"].value)
        with ZipFile(BytesIO(translated)) as archive:
            self.assertEqual(source_large_sheet, archive.read("xl/worksheets/sheet2.xml"))

    # Verifies FR-2026-09-04-01.
    def test_fast_mode_reports_all_source_level_progress_stages(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            input_root = root / "input"
            output_root = root / "output"
            input_root.mkdir()
            workbook = Workbook()
            worksheet = workbook.active
            assert worksheet is not None
            worksheet["A1"] = "Visible label"
            skipped_sheet = workbook.create_sheet("Skipped")
            skipped_sheet["A1001"] = "Skipped label"
            workbook.save(input_root / "workbook.xlsx")
            progress_bars: list[_RecordedProgress] = []

            def make_progress(total: int, label: str) -> ProgressReporter:
                progress = _RecordedProgress(total, label)
                progress_bars.append(progress)
                return progress

            result = self._run(
                input_root,
                output_root,
                _EmptyOcrProvider(),
                _RecordingReplacementProvider(replacement_text="Translated"),
                xlsx_translation_mode="fast",
                show_progress=True,
                progress_factory=make_progress,
                diagnostics_enabled=True,
            )
            diagnostic = json.loads(
                (output_root / "workbook.xlsx.diagnostics.json").read_text(encoding="utf-8")
            )

        self.assertEqual(1, result.processed_files)
        self.assertEqual(1, len(progress_bars))
        self.assertEqual(3, progress_bars[0].total)
        self.assertEqual(
            [
                "xl/worksheets/sheet1.xml replacement 1",
                "chart cache synchronization",
                "package write",
            ],
            progress_bars[0].postfixes,
        )
        self.assertEqual(3, progress_bars[0].updates)
        self.assertTrue(progress_bars[0].closed)
        skipped = [
            entry for entry in diagnostic["entries"]
            if entry["reason_code"] == "xlsx_fast_mode_worksheet_skipped"
        ]
        self.assertEqual(["Skipped"], [entry["worksheet_name"] for entry in skipped])

    # Verifies FR-2026-09-03-01 and FR-2026-09-03-05.
    def test_embedded_workbook_request_total_excludes_numeric_looking_text(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "workbook.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            assert worksheet is not None
            worksheet["A1"] = "Label"
            worksheet["A2"] = 42
            worksheet["A3"] = "123"
            workbook.save(path)

            request_total = xlsx_native_text_request_total(
                path.read_bytes(), "preserve-basic-layout-source-font"
            )
            provider = _RecordingReplacementProvider(replacement_text="translated")
            translated, _replacements = replace_xlsx_bytes(
                path.read_bytes(), provider, "en", "fr", "preserve-basic-layout-source-font"
            )

        self.assertEqual(1, request_total)
        self.assertEqual(["Label"], [request.text for request in provider.requests])
        output = load_workbook(BytesIO(translated))
        output_sheet = output.active
        assert output_sheet is not None
        self.assertEqual("translated", output_sheet["A1"].value)
        self.assertEqual("123", output_sheet["A3"].value)

    # Verifies FR-2026-08-04-07.
    def test_xlsx_drawing_layout_writes_schema_ordered_run_properties(self) -> None:
        drawing = b'''<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:sp><xdr:spPr><a:xfrm><a:ext cx="914400" cy="457200"/></a:xfrm></xdr:spPr><xdr:txBody><a:bodyPr/><a:lstStyle/><a:p><a:r><a:rPr sz="2400"><a:latin typeface="Source Sans"/></a:rPr><a:t>Original</a:t></a:r><a:endParaRPr/></a:p></xdr:txBody></xdr:sp></xdr:wsDr>'''
        updated, count = _replace_drawing(
            drawing, _RecordingReplacementProvider(replacement_text="A longer replacement"), "en", "en",
            noto_typefaces(),
            True,
        )
        self.assertEqual(1, count)
        root = ElementTree.fromstring(updated)
        paragraph = next(item for item in root.iter() if item.tag.endswith("}p"))
        children = list(paragraph)
        self.assertLess(next(index for index, item in enumerate(children) if item.tag.endswith("}r")), next(index for index, item in enumerate(children) if item.tag.endswith("}endParaRPr")))
        run_properties = next(item for item in root.iter() if item.tag.endswith("}rPr"))
        self.assertNotIn("typeface", run_properties.attrib)
        self.assertTrue(any(item.tag.endswith("}latin") for item in run_properties))

    # Verifies FR-2026-08-03-14.
    def test_xlsx_basic_layout_fits_only_explicitly_bounded_cells(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            input_root = root / "input"
            output_root = root / "output"
            input_root.mkdir()
            source = input_root / "workbook.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            self.assertIsNotNone(worksheet)
            assert worksheet is not None
            worksheet.column_dimensions["A"].width = 12
            worksheet.column_dimensions["B"].width = 12
            worksheet.row_dimensions[1].height = 30
            worksheet.row_dimensions[2].height = 30
            worksheet["A1"] = "Bounded source"
            worksheet["A1"].font = Font(name="Source Sans", sz=24, bold=True)
            worksheet["A1"].alignment = Alignment(horizontal="center", wrap_text=False)
            worksheet["A1"].protection = Protection(locked=False)
            worksheet.merge_cells("A2:B2")
            worksheet["A2"] = "Merged source"
            worksheet["C1"] = "Unbounded source"
            workbook.save(source)

            result = self._run(
                input_root,
                output_root,
                _EmptyOcrProvider(),
                _RecordingReplacementProvider(replacement_text="Long replacement text " * 30),
                document_text_layout="preserve-basic-layout",
            )

            self.assertEqual(1, result.processed_files)
            self.assertEqual(3, result.replaced_native_text_items)
            output = load_workbook(output_root / "workbook.xlsx")
            output_sheet = output.active
            self.assertIsNotNone(output_sheet)
            assert output_sheet is not None
            self.assertIn("Long replacement text", output_sheet["A1"].value)
            # FR-2026-08-04-07: XLSX has no portable embedded-font path, so the
            # bounded cell keeps its source face while using the shared fitted size.
            self.assertEqual("Source Sans", output_sheet["A1"].font.name)
            self.assertLess(output_sheet["A1"].font.sz, 24)
            self.assertTrue(output_sheet["A1"].alignment.wrap_text)
            self.assertFalse(output_sheet["A1"].alignment.shrink_to_fit)
            with ZipFile(output_root / "workbook.xlsx") as archive:
                styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
            namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            cell_xfs = styles.find(namespace + "cellXfs")
            assert cell_xfs is not None
            self.assertFalse(any(
                [child.tag for child in xf].index(namespace + "protection")
                < [child.tag for child in xf].index(namespace + "alignment")
                for xf in cell_xfs
                if namespace + "protection" in [child.tag for child in xf]
                and namespace + "alignment" in [child.tag for child in xf]
            ))
            self.assertIn("Long replacement text", output_sheet["A2"].value)
            self.assertEqual("Calibri", output_sheet["A2"].font.name)
            self.assertIn("Long replacement text", output_sheet["C1"].value)
            self.assertEqual("Calibri", output_sheet["C1"].font.name)

    # Verifies FR-2026-08-03-14.
    def test_xlsx_basic_layout_retains_structured_table_headers(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            input_root = root / "input"
            output_root = root / "output"
            input_root.mkdir()
            source = input_root / "workbook.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            self.assertIsNotNone(worksheet)
            assert worksheet is not None
            worksheet.column_dimensions["A"].width = 16
            worksheet.column_dimensions["B"].width = 16
            worksheet.row_dimensions[1].height = 24
            worksheet.row_dimensions[2].height = 24
            worksheet.append(["Synthetic Header One", "Synthetic Header Two"])
            worksheet.append(["First body value", "Second body value"])
            worksheet.add_table(Table(displayName="SyntheticTable", ref="A1:B2"))
            workbook.save(source)

            result = self._run(
                input_root,
                output_root,
                _EmptyOcrProvider(),
                _RecordingReplacementProvider(replacement_text="Replacement body value"),
                document_text_layout="preserve-basic-layout",
            )

            self.assertEqual(2, result.replaced_native_text_items)
            output = load_workbook(output_root / "workbook.xlsx")
            output_sheet = output.active
            self.assertIsNotNone(output_sheet)
            assert output_sheet is not None
            self.assertEqual("Synthetic Header One", output_sheet["A1"].value)
            self.assertEqual("Synthetic Header Two", output_sheet["B1"].value)
            self.assertEqual("Replacement body value", output_sheet["A2"].value)
            self.assertEqual("Replacement body value", output_sheet["B2"].value)
            with ZipFile(source) as source_archive, ZipFile(output_root / "workbook.xlsx") as output_archive:
                self.assertEqual(
                    source_archive.read("xl/tables/table1.xml"),
                    output_archive.read("xl/tables/table1.xml"),
                )



if __name__ == "__main__":
    unittest.main()
